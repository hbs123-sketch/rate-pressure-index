"""Verify displayed utility historical-price points against EIA Form 861 files."""

from io import BytesIO
from json import load
from pathlib import Path
from urllib.request import urlopen
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

with (ROOT / "data" / "utilities.json").open(encoding="utf-8") as stream:
    UTILITIES = load(stream)


def average_residential_price(workbook_bytes, year, utility_name, state):
    with ZipFile(BytesIO(workbook_bytes)) as archive:
        filename = f"Sales_Ult_Cust_{year}.xlsx"
        with archive.open(filename) as stream:
            worksheet = load_workbook(BytesIO(stream.read()), read_only=True, data_only=True)["States"]
            rows = [
                row
                for row in worksheet.iter_rows(min_row=4, values_only=True)
                if row[2] == utility_name and row[4] == "Bundled" and row[6] == state
            ]

    revenue = sum(float(row[9]) for row in rows if isinstance(row[9], (int, float)))
    sales = sum(float(row[10]) for row in rows if isinstance(row[10], (int, float)))
    if not revenue or not sales:
        raise AssertionError(f"No bundled residential data for {utility_name} in {state}, {year}")
    return round(revenue / sales * 100, 2)


SOURCE_CACHE = {}

for utility in UTILITIES:
    series = utility["historical_rate_series"]
    for point, source_url in zip(series["values"], series["source_file_urls"], strict=True):
        if source_url not in SOURCE_CACHE:
            with urlopen(source_url, timeout=90) as response:
                SOURCE_CACHE[source_url] = response.read()
        actual = average_residential_price(SOURCE_CACHE[source_url], point["year"], utility["eia_utility_name"], utility["eia_state_code"])
        assert actual == point["value"], (
            f"{utility['utility_name']} {point['year']}: expected {point['value']}, got {actual}"
        )
    print(f"verified EIA historical series for {utility['utility_name']}")
