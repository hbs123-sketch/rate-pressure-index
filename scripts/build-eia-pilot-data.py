"""Extract pilot utility metadata and service counties from EIA Form 861.

V1 note: EIA Form 861 service territory data is county/state based, not ZIP
based. Full national ZIP-to-utility coverage needs a geocoding and territory
overlay pipeline and is intentionally a v2 task.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

PILOT_IDS = {19876, 14006, 7140}
BASE = Path("data/.cache/f8612024")


def rows(path: Path, sheet: str):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    header = [str(cell or "").strip() for cell in next(worksheet.iter_rows(values_only=True))]
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        yield dict(zip(header, values))


def main() -> None:
    utilities = {}
    for row in rows(BASE / "Utility_Data_2024.xlsx", "States"):
        utility_id = row.get("Utility Number")
        if utility_id in PILOT_IDS:
            utilities[str(utility_id)] = {
                "eia_utility_name": row.get("Utility Name"),
                "state": row.get("State"),
                "ownership": row.get("Ownership Type"),
                "nerc_region": row.get("NERC Region"),
            }

    counties = {str(utility_id): [] for utility_id in PILOT_IDS}
    for row in rows(BASE / "Service_Territory_2024.xlsx", "Counties_States"):
        utility_id = row.get("Utility Number")
        if utility_id in PILOT_IDS:
            counties[str(utility_id)].append({"state": row.get("State"), "county": row.get("County")})

    sales = {}
    for row in rows(BASE / "Sales_Ult_Cust_2024.xlsx", "States"):
        utility_id = row.get("Utility Number")
        if utility_id in PILOT_IDS and row.get("Service Type") == "Bundled":
            sales.setdefault(str(utility_id), []).append(
                {
                    "state": row.get("State"),
                    "ba_code": row.get("BA Code"),
                    "residential_revenue_thousand_dollars": row.get("Thousand Dollars"),
                    "residential_mwh": row.get("Megawatthours"),
                    "residential_customers": row.get("Count"),
                }
            )

    output = {
        "source": "https://www.eia.gov/electricity/data/eia861/zip/f8612024.zip",
        "data_year": 2024,
        "utilities": utilities,
        "service_counties": counties,
        "bundled_sales_rows": sales,
    }
    Path("data/eia-pilot-summary.json").write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
